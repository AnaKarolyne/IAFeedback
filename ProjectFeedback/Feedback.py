import speech_recognition as sr
import pyttsx3
import os
import openpyxl
from openpyxl import Workbook
import datetime

recognizer = sr.Recognizer()
engine = pyttsx3.init()

def maquina_fala(fala):
    print(fala)
    engine.say(fala)
    engine.runAndWait()

def unknown_value():
    erro = "Não foi possível entender a sua fala. Tente novamente."
    maquina_fala(erro)
    return

def request_error(e):
    erro = "Ocorreu um erro na API de reconhecimento de fala: {e}"
    maquina_fala(erro)
    return

def mostrar_feedbacks():
    # Nome do arquivo Excel a ser lido
    arquivo_excel = "feedbacks.xlsx"

    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
    except FileNotFoundError:
        erro = "O arquivo Excel ainda não existe ou foi excluído."
        maquina_fala(erro)
        return

    # Abre a planilha padrão (Sheet1)
    sheet = workbook.active

    # Verifica se a planilha não está vazia
    if sheet.max_row <= 1:
        erro = "Não há registros de feedback no arquivo Excel."
        maquina_fala(erro)
        return

    # Itera pelas linhas do arquivo Excel e exibe os registros
    registros = "Registros de Feedback:"
    maquina_fala(registros)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        #Única string com todos os valores da linha
        linha_completa = " | ".join(map(str, row))
        print(linha_completa)

    workbook.close()

def coletar_usabilidade():
    usabilidade1 = input("Como você avaliaria a facilidade de uso do software em uma escala de 1 a 10? ")
    usabilidade2 = input("Quais aspectos específicos da interface do usuário você acha que podem ser melhorados? ")
    usabilidade3 = input("Você teve algum problema ao navegar pelo software? Se sim, por favor, descreva. ")

    # Adiciona rótulos para diferenciar cada categoria
    categorias_usabilidade = {
        "Avaliação de Facilidade de Uso": usabilidade1,
        "Aspectos a Melhorar na Interface": usabilidade2,
        "Problemas ao Navegar": usabilidade3
    }

    return categorias_usabilidade

def coletar_geral():
    geral1 = input("Em uma escala de 1 a 10, quão satisfeito você está com o software? ")
    geral2 = input("Você recomendaria este software a outros profissionais? Por quê? ")
    geral3 = input("Existe algum feedback adicional que você gostaria de compartilhar sobre sua experiência geral com o software? ")

    # Adiciona rótulos para diferenciar cada categoria
    categorias_geral = {
        "Satisfação": geral1,
        "Indicações": geral2,
        "Feedback Livre": geral3
    }

    return categorias_geral

def coletar_atualizacoes():
    atualizacoes1 = input("Como você se sente em relação às atualizações frequentes do software? ")
    atualizacoes2 = input("Há alguma atualização recente que você achou particularmente útil ou problemática? ")
    atualizacoes3 = input("Que tipo de melhorias você gostaria de ver nas próximas versões do software? ")

    # Adiciona rótulos para diferenciar cada categoria
    categorias_atualizacoes = {
        "Sentimento de progresso": atualizacoes1,
        "Pontos de atenção": atualizacoes2,
        "Recomendações": atualizacoes3
    }

    return categorias_atualizacoes

def coletar_seguranca():
    seguranca1 = input("Você se sente seguro ao usar este software em relação à segurança de seus dados? ")
    seguranca2 = input("Existe alguma preocupação específica de segurança ou privacidade que você gostaria de mencionar? ")
    seguranca3 = input("Você percebeu alguma vulnerabilidade de segurança enquanto usava o software? ")

    # Adiciona rótulos para diferenciar cada categoria
    categorias_seguranca = {
        "Proteção e Privacidade de Dados": seguranca1,
        "Segurança Falha": seguranca2,
        "Vulnerabilidades": seguranca3
    }

    return categorias_seguranca

def coletar_suporte():
    suporte1 = input("Como você avalia a qualidade do suporte técnico fornecido pela empresa? ")
    suporte2 = input("Você teve uma boa experiência com o atendimento ao cliente? Pode compartilhar detalhes? ")
    suporte3 = input("Existe algum incidente de suporte técnico que você gostaria de destacar? ")

    # Adiciona rótulos para diferenciar cada categoria
    categorias_suporte = {
        "Qualidade do Suporte": suporte1,
        "Atendimento ao cliente": suporte2,
        "Incidentes": suporte3
    }

    return categorias_suporte


def coletar_recursos():
    recursos1 = input("Quais recursos ou funcionalidades do software você considera mais úteis? ")
    recursos2 = input("Há alguma funcionalidade que você gostaria que o software tivesse, mas não tem atualmente? ")
    recursos3 = input("Você teve dificuldades em encontrar alguma função específica no software? ")

    # Adiciona rótulos para diferenciar cada categoria
    categorias_recursos = {
        "Pontos úteis": recursos1,
        "O que seria importante ter": recursos2,
        "Dificuldades": recursos3
    }

    return categorias_recursos

def coletar_desempenho():
    desempenho1 = input("O software atendeu às suas expectativas em termos de desempenho? ")
    desempenho2 = input("Você notou alguma lentidão ou travamento ao usar o software? ")
    desempenho3 = input("Em que situações o software parece mais lento para você? ")

    # Adiciona rótulos para diferenciar cada categoria
    categorias_desempenho = {
        "Desempenho": desempenho1,
        "Velocidade": desempenho2,
        "situações de Timeout": desempenho3
    }

    return categorias_desempenho

# Função para coletar feedback e armazená-lo no arquivo Excel
def coletar_feedback():

    nome_cliente = input("Qual é o seu nome? ")
    nome_produto = input(f"Qual é o nome do produto que você vai avaliar {nome_cliente}?")

    try:
        # Chamando a função coletar_usabilidade
        categorias_usabilidade = coletar_usabilidade()
        categorias_desempenho = coletar_desempenho()
        categorias_recurso = coletar_recursos()
        categorias_suporte = coletar_suporte()
        categorias_seguranca = coletar_seguranca()
        categorias_atualizacoes = coletar_atualizacoes()
        categorias_geral = coletar_geral()

        salvar_feedback(
            nome_cliente,
            nome_produto,
            **categorias_usabilidade,
            **categorias_desempenho,
            **categorias_recurso,
            **categorias_suporte,
            **categorias_seguranca,
            **categorias_atualizacoes,
            **categorias_geral)

    except sr.UnknownValueError:
        unknown_value()
        coletar_feedback()

    except sr.RequestError as e:
        request_error()
        coletar_feedback()

def salvar_feedback(nome_cliente, nome_produto, **categorias):
    nome_arquivo = "feedbacks.xlsx"

    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
    except FileNotFoundError:
        workbook = Workbook()

    sheet = workbook.active

    if "Sheet" not in workbook.sheetnames:
        sheet.title = "Feedbacks"

    cabecalhos = ["Data e Hora", "Nome do Cliente", "Nome do Produto Avaliado"] + list(categorias.keys())
    if sheet.max_row == 1:
        sheet.append(cabecalhos)

    data_hora = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    nova_linha = [data_hora, nome_cliente, nome_produto] + list(categorias.values())
    sheet.append(nova_linha)

    workbook.save(nome_arquivo)
    workbook.close()

    sucesso = "Feedback armazenado com sucesso!"
    maquina_fala(sucesso)
    principal()


# Função principal - Íncio da trilha
def trilha():
    comeco = "Bem vindo a nossa trilha de Feedback."
    coletar_feedback()

def funcoes():
    while True:
        with sr.Microphone() as source:
            comando = input("Eu fui projetada especialmente para te ouvir, entender quais são as suas críticas em relação ao nosso produto e descobrir, com a sua ajuda, como podemos fornecer um produto e atendimento melhor para você. Você gostaria de dar um feedback ou visualizar os feedbacks coletados?")
        try:
            if "dar" in comando or "novo" in comando:
                trilha()
            elif "visualizar" in comando or "Visualizar" in comando or "ver" in comando or "analisar" in comando:
                mostrar_feedbacks()
            elif "não" in comando or "Não" in comando or "nenhuma" in comando:
                negativa = "Tudo bem, espero te ver numa próxima vez e que você tenha uma experiência melhor com o nosso produto."
                principal()
            else:
                erro = "Não entendi o comando. Tente novamente."
                funcoes()
        except sr.UnknownValueError:
            unknown_value()

        except sr.RequestError as e:
            request_error(e)

def principal():
    while True:
        with sr.Microphone() as source:
            comando = input("Olá, eu sou a IA VozDoCliente, uma inteligência artificial desenvolvida especialmente para te ouvir. Vamos começar? ")
            try:
                if "sim" in comando or "começar" in comando or "Sim" in comando:
                    mestre = "Fico feliz em ajudar e a colaborar para o desenvolvimento de produtos mais assertivos."
                    funcoes()
                else:
                    mestre = "Não entendi... Talvez eu não seja a IA que você procure."
                    principal()
            except sr.UnknownValueError:
                unknown_value()

            except sr.RequestError as e:
                request_error(e)

principal()
