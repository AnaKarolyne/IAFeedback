import speech_recognition as sr
import pyttsx3
import os
import openpyxl
from openpyxl import Workbook
import datetime

recognizer = sr.Recognizer()
engine = pyttsx3.init()

# Nome do arquivo Excel com as perguntas programadas pelo gestor
arquivo_config = "perguntas_categorias_personalizadas.xlsx"

def maquina_fala(fala):
    print(fala)
    engine.say(fala)
    engine.runAndWait()

def unknown_value():
    erro = "Não foi possível entender a sua fala. Tente novamente."
    maquina_fala(erro)
    return

def request_error():
    erro = "Ocorreu um erro na API de reconhecimento de fala: {e}"
    maquina_fala(erro)
    return

def mostrar_feedbacks():
    # Nome do arquivo Excel a ser lido
    arquivo_excel = "feedbacks_personalizados"

    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
    except FileNotFoundError:
        erro = "O arquivo Excel ainda não existe ou foi excluído."
        maquina_fala(erro)
        return

    # Abre a planilha padrão (Sheet1)
    sheet = workbook.active

    # Verifica se a planilha não está vazia
    if sheet.max_row <= 1:
        erro = "Não há registros de feedback no arquivo Excel."
        maquina_fala(erro)
        return

    # Itera pelas linhas do arquivo Excel e exibe os registros
    registros = "Registros de Feedback:"
    maquina_fala(registros)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Única string com todos os valores da linha
        linha_completa = " | ".join(map(str, row))
        print(linha_completa)

    workbook.close()

def definir_categorias_perguntas(arquivo_config):
    # Verifica se o arquivo já existe no diretório atual
    try:
        workbook = openpyxl.load_workbook(arquivo_config)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.save(arquivo_config)
        workbook = openpyxl.load_workbook(arquivo_config)

    # Adicione uma planilha (worksheet) padrão
    sheet = workbook.active

    categorias_perguntas = {}

    while True:
        categoria = input("Digite o nome de uma categoria (ou 'fim' para encerrar): ")
        if categoria.lower() == 'fim':
            break

        perguntas = []
        while True:
            pergunta = input(f"Digite uma pergunta para a categoria '{categoria}' (ou 'fim' para encerrar): ")
            if pergunta.lower() == 'fim':
                break
            perguntas.append(pergunta)

        categorias_perguntas[categoria] = perguntas

    # Adiciona as novas categorias e perguntas ao arquivo Excel
    for categoria, perguntas in categorias_perguntas.items():
        for pergunta in perguntas:
            sheet.append([categoria, pergunta])

    workbook.save(arquivo_config)
    workbook.close()

def ler_categorias_perguntas(arquivo_config):
    categorias_perguntas = {}

    try:
        workbook = openpyxl.load_workbook(arquivo_config)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            categoria, pergunta = row
            if categoria not in categorias_perguntas:
                categorias_perguntas[categoria] = []
            categorias_perguntas[categoria].append(pergunta)
    except FileNotFoundError:
        print(f"O arquivo '{arquivo_config}' não existe.")
    return categorias_perguntas

def feedback_principal():
    nome_cliente = input("Qual é o seu nome? ")
    nome_produto = input(f"Qual é o nome do produto que você vai avaliar {nome_cliente}? ")
    coletar_feedback(nome_cliente, nome_produto, arquivo_config)

# Função para coletar feedback e armazená-lo no arquivo Excel
def coletar_feedback(nome_cliente, nome_produto, arquivo_config):
    try:

        # Obtém as categorias e perguntas do arquivo_config
        categorias_perguntas = ler_categorias_perguntas(arquivo_config)

        # Crie um dicionário para armazenar as respostas do usuário
        respostas = {}

        # Pergunte ao usuário as respostas para cada pergunta
        for categoria, perguntas in categorias_perguntas.items():
            respostas_categoria = {}
            for pergunta in perguntas:
                resposta = input(f"{categoria} - {pergunta} ")
                respostas_categoria[pergunta] = resposta
            respostas[categoria] = respostas_categoria

        # Chame a função salvar_feedback com os dados coletados
        salvar_feedback(nome_cliente, nome_produto, respostas, arquivo_config)

    except sr.UnknownValueError:
        unknown_value()
        coletar_feedback()

    except sr.RequestError as e:
        request_error()
        coletar_feedback()

def salvar_feedback(nome_cliente, nome_produto, respostas, arquivo_config):
    # Nome do arquivo Excel que você deseja criar ou verificar
    nome_arquivo = "feedbacks_personalizados.xlsx"

    # Verifica se o arquivo já existe no diretório atual
    if os.path.isfile(nome_arquivo):
        print(f"O arquivo '{nome_arquivo}' já existe.")
    else:
        try:
            # Crie um novo arquivo Excel (workbook)
            workbook = openpyxl.Workbook()

            # Adicione uma planilha (worksheet) padrão
            sheet = workbook.active

            # Adicione cabeçalhos dinâmicos a partir do arquivo de configuração
            categorias_perguntas = ler_categorias_perguntas(arquivo_config)
            coluna_atual = 1

            # Adiciona a coluna de data e hora
            sheet.cell(row=1, column=coluna_atual, value="Data e Hora")
            coluna_atual += 1

            # Adiciona a coluna de data e hora
            sheet.cell(row=1, column=coluna_atual, value="Nome")
            coluna_atual += 1

            # Adiciona as colunas dinâmicas para cada pergunta
            for categoria, perguntas in categorias_perguntas.items():
                for pergunta in perguntas:
                    sheet.cell(row=1, column=coluna_atual, value=f"{categoria} - {pergunta}")
                    coluna_atual += 1

            # Salve o arquivo Excel com um nome específico
            workbook.save(nome_arquivo)

            # Feche o arquivo (não é estritamente necessário, mas é uma boa prática)
            workbook.close()

            print(f"O arquivo feedbacks.xlsx foi criado com sucesso.")

        except FileNotFoundError:
            workbook = Workbook()

    # Abre o arquivo Excel existente
    workbook = openpyxl.load_workbook(nome_arquivo)

    # Selecione a planilha existente ou crie uma nova, se necessário
    if "Feedback" in workbook.sheetnames:
        sheet = workbook["Feedback"]
    else:
        sheet = workbook.create_sheet("Feedback")

    # Obtém a data e hora atual
    data_hora = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Adiciona o feedback à planilha
    nova_linha = [data_hora, nome_cliente, nome_produto]

    # Adiciona as respostas às colunas correspondentes
    for categoria, perguntas in respostas.items():
        for pergunta, resposta in perguntas.items():
            nova_linha.append(resposta)

    sheet.append(nova_linha)

    # Salva as alterações no arquivo Excel
    workbook.save(nome_arquivo)

    # Feche o arquivo (não é estritamente necessário, mas é uma boa prática)
    workbook.close()

    sucesso = "Feedback armazenado com sucesso!"
    maquina_fala(sucesso)
    principal()

# Função principal - Íncio da trilha
def trilha():
    comeco = "Bem vindo a nossa trilha de Feedback."
    maquina_fala(comeco)
    feedback_principal()

def funcoes():
    while True:
        with sr.Microphone() as source:
            comando = "Eu fui projetada especialmente para te ouvir, entender quais são as suas críticas em relação ao nosso produto e descobrir, com a sua ajuda, como podemos fornecer um produto e atendimento melhor para você. Você é o gestor ou o usuário final?"
            maquina_fala(comando)
            comando = input()
        try:
            if "gestor" in comando:
                with sr.Microphone() as source:
                    comando = "Você gostaria de adicionar perguntas e categorias novas no seu programa de feedback ou visualizar os feedbacks coletados?"
                    maquina_fala(comando)
                    comando = input()

                    try:
                        if "visualizar" in comando or "Visualizar" in comando or "ver" in comando or "analisar" in comando:
                            mostrar_feedbacks()
                        elif "adicionar" in comando:
                            definir_categorias_perguntas(arquivo_config)
                        elif "fim" in comando:
                            break
                        else:
                            erro = "Não entendi o comando. Tente novamente."
                            maquina_fala(erro)
                            funcoes()
                    except sr.UnknownValueError:
                        unknown_value()

                    except sr.RequestError as e:
                        request_error()

            else:
                with sr.Microphone() as source:
                    comando = "Você gostaria de dar um novo feedback?"
                    maquina_fala(comando)
                    comando = input()

                    try:
                        if "não" in comando:
                            erro = "Não entendi o comando. Tente novamente."
                            maquina_fala(erro)
                            funcoes()
                        elif "fim" in comando:
                            break
                        else:
                            trilha()
                    except sr.UnknownValueError:
                        unknown_value()

                    except sr.RequestError as e:
                        request_error()

            if "dar" in comando or "novo" in comando:
                trilha()
            elif "visualizar" in comando or "Visualizar" in comando or "ver" in comando or "analisar" in comando:
                mostrar_feedbacks()
            elif "não" in comando or "Não" in comando or "nenhuma" in comando:
                negativa = "Tudo bem, espero te ver numa próxima vez e que você tenha uma experiência melhor com o nosso produto."
                maquina_fala(negativa)
                principal()
            else:
                erro = "Não entendi o comando. Tente novamente."
                maquina_fala(erro)
                funcoes()
        except sr.UnknownValueError:
            unknown_value()

        except sr.RequestError as e:
            request_error()


def principal():
    while True:
        with sr.Microphone() as source:
            comando = "Olá, eu sou a IA VozDoCliente, uma inteligência artificial desenvolvida especialmente para te ouvir. Vamos começar? "
            maquina_fala(comando)
            comando = input()

            try:
                if "sim" in comando or "começar" in comando or "Sim" in comando:
                    funcoes()
                else:
                    mestre = "Não entendi... Talvez eu não seja a IA que você procure."
                    maquina_fala(mestre)
                    principal()

            except sr.UnknownValueError:
                unknown_value()

            except sr.RequestError as e:
                request_error()


principal()
